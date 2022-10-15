
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook, workbook, worksheet
import random
from tkcalendar import DateEntry
import datetime as dt
import pytz
import webbrowser
from tkinter import *
from PIL import ImageTk
from tkinter import messagebox
import os
import email_function2
import email_function

def delete():
    t1.delete(0, 'end')
    t3.delete(0, 'end')
    t4.delete(0, 'end')
    t5.delete(0, 'end')


def generator():
    name = str(t1.get())
    course = clicked.get()
    doc = str(t3.get())
    doi = str(t4.get())
    id1 = str(t5.get())

    image = Image.open('Template.jpeg')
    draw = ImageDraw.Draw(image)
    points1 = 475, 210
    points2 = 430, 320
    points3 = 659, 393
    points4 = 560, 480
    font1 = ImageFont.truetype("arial.ttf", 28)
    font2 = ImageFont.truetype("arial.ttf", 25)
    font3 = ImageFont.truetype("arial.ttf", 20)
    font4 = ImageFont.truetype("arial.ttf", 15)
    draw.text(points1, name, "black", font=font1)
    draw.text(points2, course, "black", font=font2)
    draw.text(points3, doc, "black", font=font3)
    draw.text(points4, id1, "black", font=font4)
    image.save(rf'{t1.get()}.png')
    messagebox.showinfo("Success", "Certificate generator successfully")
    image.show()


def entryrandom():
    entry_random = str(random.randint(1000000000, 9999999999))
    t5.insert(END, entry_random)


wb = load_workbook('E:\FrAgnel\Internship_stuff\Discover_Technologies\Projects\Certificate Generator\Records.xlsx')

sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    sheet.cell(row=1, column=1).value = "Certificate Number"
    sheet.cell(row=1, column=2).value = "Student Name"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Date of completion"
    sheet.cell(row=1, column=5).value = "Date of issue"


def insert():
    if (t5.get() == "" and
            t1.get() == "" and
            clicked.get() == "" and
            t3.get() == "" and
            t4.get() == ""):

        print("empty input")

    else:

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = t5.get()
        sheet.cell(row=current_row + 1, column=2).value = t1.get()
        sheet.cell(row=current_row + 1, column=3).value = clicked.get()
        sheet.cell(row=current_row + 1, column=4).value = t3.get()
        sheet.cell(row=current_row + 1, column=5).value = t4.get()

        wb.save('E:\FrAgnel\Internship_stuff\Discover_Technologies\Projects\Certificate Generator\Records.xlsx')
        messagebox.showinfo("Success", "Excel Entry Recorded")


win = Tk()
win.title("Certificate Generater")
win.geometry("550x550")
win['bg'] = "Turquoise"

options = [
    "AI-101 Python Programming",
    "AI-102 Machine Programming",
    "AI-103 Android Programming",
    "AI-104 Java Programming with Database",
    "AI-105 Blockchain Security",
    "AI-106 Autodesk-3D Printing",
    "AI-108 Embedded Systems",
    "AI-109 Solar Energy"
]

clicked = StringVar()
clicked.set("Choose")

t2 = OptionMenu(win, clicked, *options)
t2.grid(row=1, column=1, padx=20, pady=25)
t2.config(width=20)

IST = pytz.timezone('Asia/Kolkata')
date_text_var = StringVar()


def insert_today_date():
    raw = dt.datetime.now(IST)
    formatted = raw.strftime("%d/%m/%Y")
    date_text_var.set(formatted)


checkbox_today = IntVar()
today_date_checkbox = Checkbutton(win, text="Today", bg='white', variable=checkbox_today, onvalue=1, offvalue=0,
                                  command=insert_today_date)
today_date_checkbox.grid(row=3, column=2)

l1 = Label(win, text="Full Name", font=("verdana", 12, "bold"), borderwidth=5).grid(row=0, column=0, padx=20,
                                                                                    pady=25)
t1 = Entry(win, borderwidth=7, width=20, font=("verdana 10 bold"))
t1.grid(row=0, column=1, padx=20, pady=25)

l2 = Label(win, text="Course Name", font=("verdana", 12, "bold"), borderwidth=5).grid(row=1, column=0, padx=20,
                                                                                      pady=25)
# t2 = Entry(win, borderwidth=7, width=20, font=("verdana 10 bold"))
# t2.grid(row=1, column=1, padx=20, pady=25)

l3 = Label(win, text="Date of completion", font=("verdana", 12, "bold"), borderwidth=5).grid(row=2, column=0, padx=20,
                                                                                             pady=25)
t3 = DateEntry(win, width=25, background="magenta3", date_pattern='dd/mm/y', foreground="white", bd=2)
t3.grid(row=2, column=1, padx=20, pady=25)
# t3 = Entry(win, borderwidth=7, width=20, font=("verdana 10 bold"))
# t3.grid(row=2, column=1, padx=20, pady=25)

l4 = Label(win, text="Date of issual", font=("verdana", 12, "bold"), borderwidth=5).grid(row=3, column=0, padx=20,
                                                                                         pady=25)
t4 = Entry(win, borderwidth=7, width=20, font=("verdana 10 bold"), textvariable=date_text_var)
t4['textvariable'] = date_text_var
t4.grid(row=3, column=1, padx=20, pady=25, columnspan=1)

# t4 = Label(win, text=f"{dt.datetime.now():%m/%d/%y}", fg="white", bg="black", width=15, font=("Arial", 14))
# t4.grid(row=3, column=1, padx=20, pady=25)

l5 = Label(win, text="Unique ID", font=("verdana", 12, "bold"), borderwidth=5).grid(row=4, column=0, padx=20,
                                                                                    pady=25)
t5 = Entry(win, borderwidth=7, width=20, font=("verdana 10 bold"))
t5.grid(row=4, column=1, padx=20, pady=25)

generate = Button(win, text="Generate", width=17, borderwidth=5, font=("verdana 8 bold"), command=generator).place(x=35,
                                                                                                                   y=420)
clear = Button(win, text="Clear", width=17, borderwidth=5, font=("verdana 8 bold"), command=delete).place(x=235, y=420)

id1 = Button(win, text="Click for Unique ID", width=17, borderwidth=5, font=("verdana 8 bold"),
             command=entryrandom).place(x=35, y=470)

excel()

excel1 = Button(win, text="Generate Excel Sheet", width=17, borderwidth=5, font=("verdana 8 bold"),
                command=insert).place(x=235, y=470)



win.mainloop()
